#!/usr/bin/env python3
"""
MRTG Bandwidth Report Generator
================================
Extracts Maximum Usage (Mbps) from MRTG/Cacti graph PDFs and generates
a Bandwidth Report (MAX) xlsx file from a template.

ACCURACY NOTE:
  This tool uses OCR (Tesseract) to read text from graph images in PDFs.
  OCR accuracy on MRTG graphs is typically ~80-90%. Some values may be
  misread, especially:
    - Unit letters (M/G/k) can be missed or misread
    - Decimal points can be confused with commas
    - Graph titles with OCR artifacts may not match patterns
  ALWAYS manually verify the output xlsx against the source PDF,
  especially for critical or unusual values.

Requirements:
  Python packages: pip install openpyxl pdf2image pytesseract Pillow
  System packages: Tesseract OCR, Poppler (pdftoppm)

Usage:
  python mrtg_bandwidth_report.py                 # Launch GUI
  python mrtg_bandwidth_report.py --cli --pdf <input.pdf> --template <template.xlsx> --date "26 March 2026"

Installation (one-liner):
  macOS:   brew install tesseract poppler && pip install openpyxl pdf2image pytesseract Pillow
  Ubuntu:  sudo apt install -y tesseract-ocr poppler-utils && pip install openpyxl pdf2image pytesseract Pillow
  Windows: choco install tesseract poppler && pip install openpyxl pdf2image pytesseract Pillow
"""

import re
import os
import sys
import json
import shutil
import logging
import argparse
from datetime import datetime
from difflib import SequenceMatcher

# ---------------------------------------------------------------------------
# Third-party imports (checked at startup)
# ---------------------------------------------------------------------------
MISSING_DEPS = []
try:
    from openpyxl import load_workbook
    from openpyxl.styles import PatternFill, Font, Border, Side
except ImportError:
    MISSING_DEPS.append("openpyxl")
try:
    from pdf2image import convert_from_path
except ImportError:
    MISSING_DEPS.append("pdf2image")
try:
    import pytesseract
    from PIL import Image
except ImportError:
    MISSING_DEPS.append("pytesseract Pillow")

logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s", datefmt="%H:%M:%S")
log = logging.getLogger("mrtg")


# =====================================================================
# SECTION 1 — GRAPH CLIENT NAME → SPREADSHEET ROW MAPPING
# =====================================================================
# The graph title from OCR typically looks like:
#   "1-IPT-BSCPLC-COX-03 - Coronet-IPT - HundredGigE0/0/"
#   "2-IPT-BSCPLC-DHK-CORE-03 - Velocity-DhakaColo - Bundle-Ether655"
#
# We extract the CLIENT KEYWORD (middle part between " - " delimiters)
# and match against these patterns.
#
# Each entry: (regex_pattern, excel_cell, description)
# Pattern is matched against the CLIENT KEYWORD (case-insensitive).
# First match wins — put more specific patterns first.
# =====================================================================

GRAPH_TO_ROW_MAP = [
    # ---- IIG Clients (rows 4-28) ----
    (r"BDHUB.*[IL]IG|BDHUB-LIG|BDHUB-15G-IIG", "E4", "BDHUB DHK IIG"),
    (r"Equitel|EQUITEL", "E5", "Equitel DHK"),
    (r"Skytel.*DC|Skytel.*PRIMARY|Skytel.*IIG.*PRI", "E6", "Skytel Primary"),
    (r"Skytel.*TEJ|Skytel.*SECONDARY|Skytel.*IIG.*SEC", "E7", "Skytel Secondary"),
    (r"PEEREX.*TEJ\b|PEEREX-TEJ", "E8", "Peerex DHK"),
    (r"PEEREX.*9\.?5G|PEEREX-9", "E9", "Peerex Cox-9500"),
    (r"PEEREX.*COX.*[0@][2z]|PEEREX-COX-[0@][2z]", "E10", "Peerex Cox-3432"),
    (r"F.H.*KKT.*BE|FAH.*KKT|F@H.*KKT", "E11", "F@Home KKT"),
    (r"NOVOCOM|Novocom", "E12", "NOVOCOM DHK"),
    (r"WINDSTREAM.*IIG|Windstream.*IIG", "E13", "Windstream COX IIG"),
    (r"Velocity.*Tej|VELOCITY.*TEJ", "E14", "Velocity DHK-Tej-Primary"),
    (r"Velocity.*DHK|Velocity.*DhakaColo|Velocity.*Khaza|VELOCITY.*DHK", "E15", "Velocity DHK-Khaza-Sec"),
    (r"Virgo|VIRGO", "E16", "Virgo DHK"),
    (r"DELTA.IPT(?!.*LD)|DELTA-COX\b(?!.*LD)|Delta-COX\b(?!.*LD)", "E17", "Delta COX IIG"),
    (r"Exabyte-IPT|Exabyte.*IPT.*Bundle-Ether172", "E18", "Exabyte COX IIG"),
    (r"Coronet-IPT|Coronet.*IPT|CORONET.*IPT|C.?RONET.*IPT", "E19", "Coronet COX IIG"),
    (r"INTRAGLOBE.*IPT|Intraglobe.*IPT", "E20", "Intraglobe KKT IIG"),
    (r"GMax-IPT|GMAX.*IPT|GMax.*IPT", "E21", "Green Max COX IIG"),
    (r"BD.?LINK.*(?:IIG|DC)|BDLINK.*DC", "E22", "BD-LINK DHK"),
    (r"ADNGateway.*SEC|ADN.*Gateway.*SEC|ADN.*GW.*SEC|ADN.*TEJ", "E24", "ADN-GW DHK-Secondary"),
    (r"ADNGateway|ADN.*Gateway", "E23", "ADN-GW DHK-Primary"),
    (r"REGO.COX.IIG|REGO_COX_IIG|REGO.*COX.*IIG", "E25", "Rego COX IIG"),
    (r"REGO-IIG|REGO.*IIG(?!.*COX)", "E26", "Rego KKT IIG"),
    (r"GFCL-IPT|GFCL.*IPT", "E27", "GFCL COX IIG"),
    (r"MAXHUB.*COX|BE-MAXHUB|MAX.?HUB", "E28", "Max Hub Ltd COX"),

    # ---- ISP Clients (rows 31-51) ----
    # NOTE: E32 (SEC) must come before E31 (Primary) — first match wins
    (r"ADN.*DHK.*SEC|ADN.*DhakaColo.*SEC", "E32", "ADN DHK-Secondary"),
    (r"ADN.*DHK.*ISP|ADN-DhakaColo|ADN.*DhakaColo|ADN.*DC(?!.*LD)", "E31", "ADN DHK-Primary"),
    (r"TELET.LK.*MOG|Teletalk.*PRI.*DHK|TELETALK.*PRI.*DHK|Teletalk.*MOG", "E33", "Teletalk DHK-Primary"),
    (r"Telet.lk.*CTG.*Sec|TELET.LK.*CTG.*Sec|Teltatalk.*CTG.*Sec", "E35", "Teletalk CTG-Secondary"),
    (r"Teletalk.*PRI.*CTG|TELETALK.*PRI.*CTG|Teletalk.*CTG(?!.*Sec)", "E34", "Teletalk CTG-Primary"),
    (r"COL.CTG.Pri|COL.*CTG.*Pri|COL.CTG.BE|COL.*CTG.*BE", "E36", "COL CTG-Primary"),
    (r"COL.CTG.SEC|COL.*CTG.*SEC|COL.CTG.Sec", "E37", "COL CTG-Secondary"),
    (r"COXLINKT|COX.?LINK.*COX", "E38", "COX-Link COX"),
    (r"SSOnline.*Cloud(?!flare)|SS.?Online.*Cloud(?!flare)|SSOnline.*DC|SS.?Online.*DC", "E39", "SS Online DHK-Primary"),
    (r"BDREN.*PRI|BDREN.*DHK.*03.*TO.*CGS|BDREN-DHK.*CGS", "E41", "BDREN DHK-Primary"),
    (r"BDREN.*SEC(?!.*Equinix)", "E42", "BDREN DHK-Secondary"),
    (r"BDCCL|BD.?CCL", "E43", "BDCCL DHK"),
    (r"Link3.*Dhaka.*Colo|LINK3.*DC|Link3.*DC", "E44", "Link3 DHK-Primary-DhakColo"),
    (r"Link3.*Tej|LINK3.*TEJ", "E45", "Link3 DHK-Secondary-Tejgaon"),
    (r"Dhaka.?Link.*Pri|DHAKA.?LINK.*PRI", "E46", "Dhaka-Link DHK-Primary"),
    (r"Dhaka.?Link.*Sec|DHAKA.?LINK.*SEC", "E47", "Dhaka-Link DHK-Secondary"),
    (r"BDREN.*Equinix|BDREN.*CC.*Equinix", "E48", "BDREN DHK (Equinix)"),
    (r"Race.?Online|RACE.?ONLINE", "E49", "Race Online Ltd"),
    (r"Telnet.*ICT|TELNET.*ICT|Telnet.*DC|Telnet-DC", "E50", "Telnet Dhk-Colo"),
    # ---- Cache (rows 54-55) ----
    (r"Exabyte.*Cloudflare.*TEJ", "E54", "Exabyte Cache TEJ"),
    (r"EDGENEXT.*CLOUD|BE-EDGENEXT", "E54", "Exabyte Cache (EDGENEXT)"),
    (r"Exabyte.*Cloudflare.*DC", "E54", "Exabyte Cache DC"),
    (r"SS.?Online.*Cache|SSOnline.*Cloudflare|SS.?Online.*CDN", "E55", "SS Online Cache DHK"),

    # ---- LD Clients (rows 58-67) ----
    (r"DELTA.LD|Delta.LD", "E58", "Delta COX LD"),
    (r"Intraglobe.LD|INTRAGLOBE.LD", "E59", "Intraglobe KKT LD"),
    (r"Coronet.LD|CORONET.LD|C.?RONET.LD", "E60", "Coronet COX LD"),
    (r"GFCL.LD|GFCL.*LD", "E61", "GFCL COX LD"),
    (r"BDHUB.LD(?!.*15G)|BDHUB.*LD.*313", "E62", "BDHUB COX LD"),
    (r"GMAX.LD|GMax.LD|Green.?Max.*LD", "E63", "Green Max COX LD"),
    (r"Exabyte.LD|EXABYTE.LD|EXABYTE.*COX.*LD", "E64", "Exabyte COX LD"),
    (r"Windstream.LD|WINDSTREAM.LD", "E65", "Windstream KKT LD"),
    (r"COL.COX.LD|C0L.COX.LD", "E66", "COL COX LD"),
    (r"SS.?Online.LD|SSONLINE.LD", "E67", "SS Online COX LD"),
    (r"BDHUB.LD.15G|BDHUB.*LD.*15G", "E62", "BDHUB COX LD (15G)"),
    (r"Delta.LD.*5000.*Cox|Delta-LD-5000", "E58", "Delta COX LD (Cox's Bazar)"),
]

# Additional broad fallbacks — used when the full title is tried.
# NOTE: Do NOT add entries already covered by GRAPH_TO_ROW_MAP; they will
# never fire since GRAPH_TO_ROW_MAP is checked first.
FALLBACK_MAP = [
    (r"GREENMAX.*COX(?!.*LD)", "E21", "Green Max COX IIG (GREENMAX)"),
    (r"Windstrem.*IPT|WINDSTREAM.*IPT(?!.*IIG)", "E13", "Windstream COX (typo)"),
    (r"PEEREX.*DHKCOLO", "E8", "Peerex DHK (DHKCOLO)"),
    (r"DELTA-IPT.*Ether193", "E17", "Delta COX IIG (IPT/193)"),
    (r"Telt?at?alk.*CTG.*Sec|Teltatalk.*Sec", "E35", "Teletalk CTG-Sec (OCR)"),
    (r"TELET\w+.*MOG|Teletalk.*MOG", "E33", "Teletalk DHK (MOG)"),
    (r"BDLINK|BD.LINK(?!.*LD)", "E22", "BD-LINK DHK (simple)"),
    (r"Windstrem.*I.PT|Windstrem.*\[PT", "E13", "Windstream COX IIG (OCR)"),
]


# =====================================================================
# SECTION 2 — OCR + PARSING ENGINE
# =====================================================================

def pdf_to_images(pdf_path: str, dpi: int = 250) -> list:
    """Convert PDF pages to PIL Image list."""
    log.info(f"Converting PDF to images at {dpi} DPI ...")
    images = convert_from_path(pdf_path, dpi=dpi)
    log.info(f"  -> {len(images)} pages converted.")
    return images


def ocr_full_page(page_img: Image.Image) -> str:
    """Run Tesseract OCR on a full page image."""
    try:
        return pytesseract.image_to_string(page_img, config="--psm 6")
    except Exception as e:
        log.warning(f"OCR failed: {e}")
        return ""


def extract_client_keyword(title: str) -> str:
    """
    Extract the client keyword from a graph title.
    Input:  "1-IPT-BSCPLC-COX-03 - Coronet-IPT - HundredGigE0/0/"
    Output: "Coronet-IPT"

    The client name is typically the 2nd segment when split by " - ".
    For two-part titles ("ADN-DhakaColo - TenGigE0/0/0"), the client is
    in part[0] since part[1] is the interface — detect this and return part[0].
    """
    parts = re.split(r"\s+-\s+", title)
    if len(parts) >= 3:
        return parts[1].strip()
    if len(parts) == 2:
        # If the second part looks like an interface name, the client is in part[0]
        if re.search(r"(Bundle-Ether|TenGigE|HundredGigE|FortyGigE|GigE\d|GigabitEthernet|Ethernet\d|Ether\d|(?:Gi|Te|Fo|Hu)\d+/)", parts[1], re.I):
            return parts[0].strip()
        return parts[1].strip()
    return title.strip()


def parse_graphs_from_text(text: str) -> list:
    """
    Parse full-page OCR text to extract multiple graph data blocks.

    Each MRTG graph produces text roughly like:
      <title line with interface name>
      ...graph area noise...
      Inbound   Current: X   Average: Y   Maximum: Z
      Outbound  Current: X   Average: Y   Maximum: Z
      Total: ...

    Returns list of dicts with keys: title, inbound_max, outbound_max
    """
    graphs = []
    lines = text.split("\n")

    # Strategy: find title lines (containing interface names like Bundle-Ether, TenGigE, HundredGigE)
    # then find the Inbound/Outbound stats that follow each title
    title_indices = []
    for i, line in enumerate(lines):
        if re.search(r"(Bundle-Ether|TenGigE|HundredGigE|FortyGigE|TwentyFiveGig|GigE\d|GigabitEthernet|Ethernet\d|Ether\d{2,}|(?:^|\s)(?:Gi|Te|Fo|Hu)\d+/)", line, re.I):
            # Make sure it looks like a title (has known device/client marker or prefix pattern)
            if re.search(r"(BSCPLC|IPT|IPBW|LD|CORE|DHK|COX|KKT|CTG|TEJ|BDREN|BDCCL|TELETALK|CORONET|DELTA|LINK3|PEEREX|EXABYTE|ADN|BDHUB|REGO|GFCL|MAXHUB|NOVOCOM|EQUITEL|VIRGO|VELOCITY|WINDSTREAM|INTRAGLOBE|GMAX|EDGENEXT|SSONLINE|RACE|TELNET|COXLINK|BDLINK)", line, re.I) or re.search(r"\d-\w+-\w+", line) or re.search(r"\w+\s+-\s+\w+\s+-\s+", line):
                title_indices.append(i)

    for ti_idx, ti in enumerate(title_indices):
        title = lines[ti].strip()

        # Search for Inbound/Outbound Maximum in lines after this title
        # (up to the next title or 50 lines, whichever comes first)
        next_ti = title_indices[ti_idx + 1] if ti_idx + 1 < len(title_indices) else len(lines)
        search_end = min(ti + 50, next_ti)
        search_block = "\n".join(lines[ti:search_end])

        in_max, in_unitless = _extract_maximum(search_block, "Inbound")
        out_max, out_unitless = _extract_maximum(search_block, "Outbound")

        graphs.append({
            "title": title,
            "inbound_max": in_max,
            "outbound_max": out_max,
            "in_suspect": in_unitless,
            "out_suspect": out_unitless,
            "suspect": (in_unitless or out_unitless),
        })

    return graphs


def _direction_pattern(direction: str) -> str:
    """Build an OCR-tolerant regex for 'Inbound' or 'Outbound'."""
    if direction.lower() == "inbound":
        return r"\b[IiLl1][Nn]\s*[Bb][Oo][Uu][Nn][Dd]"
    return r"\b[Oo0Qq][Uu][Tt]\s*[Bb][Oo][Uu][Nn][Dd]"


def _extract_maximum(text_block: str, direction: str) -> tuple:
    """
    Extract the Maximum value for Inbound or Outbound from a text block.
    Returns (value_in_mbps, is_unitless) or (None, False) if not found.
    """
    for line in text_block.split("\n"):
        if not re.search(_direction_pattern(direction), line, re.I):
            continue
        pat = r"Max\w*[:\s]+([-\d.,]+)\s*([GgMmKk]?)(?:bps)?\b"
        match = re.search(pat, line, re.I)
        if match:
            val_str = match.group(1).strip()
            unit = match.group(2).strip()
            if "nan" in val_str.lower():
                return None, False
            try:
                val = float(val_str.replace(",", ""))
            except ValueError:
                continue
            return convert_to_mbps(val, unit, val_str), (unit == "")

    lines = text_block.split("\n")
    for i, line in enumerate(lines):
        if re.search(_direction_pattern(direction), line, re.I):
            combined = line
            if i + 1 < len(lines):
                combined += " " + lines[i + 1]
            if i + 2 < len(lines):
                combined += " " + lines[i + 2]
            pat = r"Max\w*[:\s]+([-\d.,]+)\s*([GgMmKk]?)(?:bps)?\b"
            match = re.search(pat, combined, re.I)
            if match:
                val_str = match.group(1).strip()
                unit = match.group(2).strip()
                if "nan" in val_str.lower():
                    return None, False
                try:
                    val = float(val_str.replace(",", ""))
                except ValueError:
                    continue
                return convert_to_mbps(val, unit, val_str), (unit == "")

    return None, False


def _fix_ocr_dropped_decimal(val_str: str, value: float, unit: str) -> float:
    """
    Detect and fix OCR-dropped decimal points in parsed values.

    MRTG graphs display values like "909.11 M" or "16.61 G". When OCR drops
    the decimal point, we get "90911 M" or "1661 G" — integers that are
    unreasonably large for the unit.

    Heuristic: if the original string has no decimal point and the value
    with unit produces an implausibly large Mbps result, try reinserting
    the decimal at positions that yield a plausible MRTG reading.
    """
    if "." in val_str:
        return value  # decimal is present, no fix needed

    u = unit.upper() if unit else ""
    digits = val_str.lstrip("-").replace(",", "")

    # Only fix values with 4+ digits for M, or 3+ digits for G
    # (MRTG rarely shows values > 999 M or 99 G without a decimal)
    if u == "M" and len(digits) >= 4:
        # Try inserting decimal 2 places from the right: 90911 -> 909.11
        # Also try 1 place: 90911 -> 9091.1 (less likely)
        for pos in (2, 1, 3):
            if pos < len(digits):
                candidate = float(digits[:-pos] + "." + digits[-pos:])
                if candidate <= 50000:  # 50 Gbps is a reasonable single-link max
                    log.info(f"    [OCR-FIX] Dropped decimal: {value} {u} -> {candidate} {u}")
                    return candidate
    elif u == "G" and len(digits) >= 3:
        for pos in (2, 1):
            if pos < len(digits):
                candidate = float(digits[:-pos] + "." + digits[-pos:])
                if candidate <= 100:  # 100 Gbps is a reasonable single-link max
                    log.info(f"    [OCR-FIX] Dropped decimal: {value} {u} -> {candidate} {u}")
                    return candidate

    return value


def convert_to_mbps(value: float, unit: str, val_str: str = "") -> float:
    """
    Convert MRTG value+unit to Megabits per second.

    MRTG graph units:
      G  = Gbps  -> multiply by 1000
      M  = Mbps  -> as-is
      k  = kbps  -> divide by 1000
      (none) = bps -> divide by 1,000,000
    """
    # Fix OCR-dropped decimals before converting
    if val_str:
        value = _fix_ocr_dropped_decimal(val_str, value, unit)

    u = unit.upper() if unit else ""
    if u == "G":
        return round(value * 1000, 2)
    elif u == "M":
        return round(value, 2)
    elif u == "K":
        return round(value / 1000, 4)
    else:
        # No unit captured — treat as raw bps per user policy.
        # MRTG unit-less values are in bits/second; convert to Mbps.
        return round(value / 1_000_000, 6)


# Curated fuzzy token lookup for common OCR-garbled titles
_FUZZY_TOKEN_MAP = [
    # IIG Clients
    ({"BDHUB", "LIG"},             "E4",  "BDHUB DHK IIG (fuzzy)"),
    ({"BDHUB", "IIG"},             "E4",  "BDHUB DHK IIG (fuzzy)"),
    ({"EQUITEL"},                   "E5",  "Equitel DHK (fuzzy)"),
    ({"SKYTEL", "PRI"},            "E6",  "Skytel Primary (fuzzy)"),
    ({"SKYTEL", "SEC"},            "E7",  "Skytel Secondary (fuzzy)"),
    ({"PEEREX", "TEJ"},            "E8",  "Peerex DHK (fuzzy)"),
    ({"PEEREX", "9500"},           "E9",  "Peerex Cox-9500 (fuzzy)"),
    ({"PEEREX", "COX", "02"},      "E10", "Peerex Cox-3432 (fuzzy)"),
    ({"PEEREX", "COX", "0"},       "E10", "Peerex Cox-3432 (fuzzy-@)"),
    ({"FAHOME", "KKT"},            "E11", "F@Home KKT (fuzzy)"),
    ({"NOVOCOM"},                   "E12", "NOVOCOM DHK (fuzzy)"),
    ({"WINDSTREM", "IPT"},         "E13", "Windstream COX IIG (fuzzy)"),
    ({"WINDSTREM", "IIPT"},        "E13", "Windstream COX IIG (fuzzy-[)"),
    ({"WINDSTREAM", "IPT"},        "E13", "Windstream COX IIG (fuzzy)"),
    ({"VELOCITY", "TEJ"},          "E14", "Velocity DHK-Tej (fuzzy)"),
    ({"VELOCITY", "DHK"},          "E15", "Velocity DHK-Khaza (fuzzy)"),
    ({"VIRGO"},                     "E16", "Virgo DHK (fuzzy)"),
    ({"DELTA", "IPT"},             "E17", "Delta COX IIG (fuzzy)"),
    ({"EXABYTE", "IPT"},           "E18", "Exabyte COX IIG (fuzzy)"),
    ({"CORONET", "IPT"},           "E19", "Coronet COX IIG (fuzzy)"),
    ({"INTRAGLOBE", "IPT"},        "E20", "Intraglobe KKT IIG (fuzzy)"),
    ({"GMAX", "IPT"},              "E21", "Green Max COX IIG (fuzzy)"),
    ({"BDLINK"},                    "E22", "BD-LINK DHK (fuzzy)"),
    ({"ADN", "GATEWAY", "SEC"},    "E24", "ADN-GW DHK-Secondary (fuzzy)"),
    ({"ADN", "TEJ"},               "E24", "ADN-GW DHK-Secondary (fuzzy)"),
    ({"ADN", "GATEWAY"},           "E23", "ADN-GW DHK-Primary (fuzzy)"),
    ({"REGO", "COX", "IIG"},       "E25", "Rego COX IIG (fuzzy)"),
    ({"REGO", "IIG"},              "E26", "Rego KKT IIG (fuzzy)"),
    ({"GFCL", "IPT"},             "E27", "GFCL COX IIG (fuzzy)"),
    ({"MAXHUB"},                    "E28", "Max Hub Ltd COX (fuzzy)"),
    # ISP Clients
    ({"ADN", "DHAKACOLO", "SEC"},  "E32", "ADN DHK-Secondary (fuzzy)"),
    ({"ADN", "DHAKACOLO"},         "E31", "ADN DHK-Primary (fuzzy)"),
    ({"TELETALK", "PRI", "DHK"},   "E33", "Teletalk DHK-Primary (fuzzy)"),
    ({"TELETALK", "CTG", "PRI"},   "E34", "Teletalk CTG-Primary (fuzzy)"),
    ({"TELETALK", "CTG", "SEC"},   "E35", "Teletalk CTG-Secondary (fuzzy)"),
    ({"COL", "CTG", "PRI"},        "E36", "COL CTG-Primary (fuzzy)"),
    ({"COL", "CTG", "SEC"},        "E37", "COL CTG-Secondary (fuzzy)"),
    ({"COXLINK"},                   "E38", "COX-Link COX (fuzzy)"),
    ({"SSONLINE", "CLOUD"},        "E39", "SS Online DHK-Primary (fuzzy)"),
    ({"BDREN", "PRI"},             "E41", "BDREN DHK-Primary (fuzzy)"),
    ({"BDREN", "SEC"},             "E42", "BDREN DHK-Secondary (fuzzy)"),
    ({"BDCCL"},                     "E43", "BDCCL DHK (fuzzy)"),
    ({"LINK3", "DC"},              "E44", "Link3 DHK-Primary (fuzzy)"),
    ({"LINK3", "TEJ"},             "E45", "Link3 DHK-Secondary (fuzzy)"),
    ({"DHAKALINK", "PRI"},         "E46", "Dhaka-Link Primary (fuzzy)"),
    ({"DHAKALINK", "SEC"},         "E47", "Dhaka-Link Secondary (fuzzy)"),
    ({"BDREN", "EQUINIX"},         "E48", "BDREN DHK Equinix (fuzzy)"),
    ({"RACEONLINE"},                "E49", "Race Online Ltd (fuzzy)"),
    ({"TELNET"},                    "E50", "Telnet Dhk-Colo (fuzzy)"),
    # Cache
    ({"EXABYTE", "CLOUDFLARE"},    "E54", "Exabyte Cache (fuzzy)"),
    ({"EDGENEXT"},                  "E54", "Exabyte Cache EDGENEXT (fuzzy)"),
    ({"SSONLINE", "CACHE"},        "E55", "SS Online Cache (fuzzy)"),
    # LD Clients
    ({"DELTA", "LD"},              "E58", "Delta COX LD (fuzzy)"),
    ({"INTRAGLOBE", "LD"},         "E59", "Intraglobe KKT LD (fuzzy)"),
    ({"CORONET", "LD"},            "E60", "Coronet COX LD (fuzzy)"),
    ({"GFCL", "LD"},               "E61", "GFCL COX LD (fuzzy)"),
    ({"BDHUB", "LD"},              "E62", "BDHUB COX LD (fuzzy)"),
    ({"GMAX", "LD"},               "E63", "Green Max COX LD (fuzzy)"),
    ({"EXABYTE", "LD"},            "E64", "Exabyte COX LD (fuzzy)"),
    ({"WINDSTREAM", "LD"},         "E65", "Windstream KKT LD (fuzzy)"),
    ({"COL", "LD"},                "E66", "COL COX LD (fuzzy)"),
    ({"SSONLINE", "LD"},           "E67", "SS Online COX LD (fuzzy)"),
]

def _fuzzy_match(title: str) -> tuple:
    """Token-overlap fuzzy matching as last resort."""
    title_tokens = set(re.findall(r"[A-Z0-9]+", title.upper()))
    best_score = 0.0
    best_row = None
    best_desc = None
    for required_tokens, row, desc in _FUZZY_TOKEN_MAP:
        if not required_tokens:
            continue
        overlap = len(required_tokens & title_tokens)
        score = overlap / len(required_tokens)
        if score >= 0.8 and score > best_score:
            best_score = score
            best_row = row
            best_desc = desc
    return best_row, best_desc


def match_graph_to_row(title: str) -> tuple:
    """
    Given a full graph title, find which spreadsheet row it maps to.
    Returns (cell_ref, description) or (None, None).

    Strategy:
    1. Extract the client keyword from the title
    2. Match against GRAPH_TO_ROW_MAP using the client keyword
    3. If no match, try matching full title against GRAPH_TO_ROW_MAP
    4. If still no match, try FALLBACK_MAP against full title
    """
    if not title:
        return None, None

    # Clean OCR artifacts — common character substitutions
    clean_title = re.sub(r"[|!]", "l", title)
    clean_title = re.sub(r"@", "0", clean_title)   # @ → 0 (common OCR)
    clean_title = re.sub(r"\[", "I", clean_title)   # [ → I (common OCR)
    clean_title = re.sub(r"\]", ")", clean_title)   # ] → ) (common OCR)
    clean_title = re.sub(r"\{", "(", clean_title)   # { → ( (common OCR)
    clean_title = re.sub(r"\}", ")", clean_title)   # } → ) (common OCR)
    clean_title = re.sub(r"\s+", " ", clean_title)

    # Extract client keyword
    client = extract_client_keyword(clean_title)

    # Try matching client keyword
    for pattern, row, desc in GRAPH_TO_ROW_MAP:
        if re.search(pattern, client, re.I):
            return row, desc

    # Try matching full title
    for pattern, row, desc in GRAPH_TO_ROW_MAP:
        if re.search(pattern, clean_title, re.I):
            return row, desc

    # Try fallback patterns on full title
    for pattern, row, desc in FALLBACK_MAP:
        if re.search(pattern, clean_title, re.I):
            return row, desc

    # Last resort: fuzzy token-overlap matching
    row_ref, desc = _fuzzy_match(clean_title)
    if row_ref:
        return row_ref, desc

    return None, None


# =====================================================================
# SECTION 3 — FULL EXTRACTION PIPELINE
# =====================================================================

# Cells where multiple graphs should be SUMMED instead of taking the max.
# E54 = Cache (Exabyte TEJ + DC + EDGENEXT all contribute to one total).
_SUM_CELLS = {"E54"}


def extract_all_graphs(pdf_path: str, dpi: int = 250, progress_cb=None,
                       warn_duplicates: bool = False, collect_raw_ocr: bool = False) -> dict:
    """
    Main pipeline: PDF → images → OCR → parsed stats → mapped to rows.

    Args:
      warn_duplicates: if True, log a WARNING whenever a row would be overwritten
                       by a higher-value graph (shows both old and new values).
      collect_raw_ocr: if True, include per-page raw OCR text in the returned dict
                       under key "raw_ocr_pages" (list of {"page": N, "text": "..."}).

    Returns dict with keys:
      results: { "E4": {"mbps": 13550, "title": "...", "desc": "..."}, ... }
      unmatched: list of unmatched graph dicts
      could_not_open: count of "Could not open!" entries
      all_graphs: all parsed graph info for debugging
      total_pages: number of pages
      raw_ocr_pages: (only when collect_raw_ocr=True) list of per-page OCR text dicts
    """
    images = pdf_to_images(pdf_path, dpi)
    total_pages = len(images)

    results = {}
    unmatched = []
    could_not_open = 0
    all_graphs = []
    raw_ocr_pages = []

    for page_idx, page_img in enumerate(images):
        page_num = page_idx + 1
        if progress_cb:
            progress_cb(page_num, total_pages, f"Processing page {page_num}/{total_pages}")
        log.info(f"Processing page {page_num}/{total_pages} ...")

        # OCR the full page
        full_text = ocr_full_page(page_img)
        if collect_raw_ocr:
            raw_ocr_pages.append({"page": page_num, "text": full_text})

        # Count "Could not open!" entries
        cno_count = len(re.findall(r"Could not open", full_text, re.I))
        could_not_open += cno_count

        # Parse graph blocks from the text
        graphs = parse_graphs_from_text(full_text)
        log.info(f"  Found {len(graphs)} graph(s) on page {page_num}")

        for g in graphs:
            in_max = g["inbound_max"] if g["inbound_max"] is not None else 0.0
            out_max = g["outbound_max"] if g["outbound_max"] is not None else 0.0
            extraction_failed = (g["inbound_max"] is None and g["outbound_max"] is None)
            max_mbps = max(in_max, out_max)

            row_ref, desc = match_graph_to_row(g["title"])

            info = {
                "page": page_num,
                "title": g["title"],
                "inbound_max": in_max,
                "outbound_max": out_max,
                "max_mbps": max_mbps,
                "row_ref": row_ref,
                "desc": desc,
                "extraction_failed": extraction_failed,
            }
            all_graphs.append(info)

            if extraction_failed and row_ref:
                log.warning(f"  EXTRACTION FAILED: {desc} -> {row_ref} — could not parse Inbound/Outbound values from OCR text")

            if row_ref:
                # Keep the larger value if duplicate (skip zero-value extraction failures)
                if extraction_failed and row_ref in results:
                    log.info(f"  SKIP (extraction failed, keeping existing): {desc} -> {row_ref}")
                elif row_ref in _SUM_CELLS and row_ref in results and not extraction_failed:
                    # Accumulate (SUM) for cells that aggregate multiple graphs
                    results[row_ref]["mbps"] += max_mbps
                    results[row_ref]["in_mbps"] += in_max
                    results[row_ref]["out_mbps"] += out_max
                    results[row_ref]["title"] += f" + {g['title']}"
                    results[row_ref]["desc"] += f" + {desc}"
                elif row_ref not in results or max_mbps > results[row_ref]["mbps"]:
                    if warn_duplicates and row_ref in results and not extraction_failed:
                        prev = results[row_ref]
                        log.warning(
                            f"  DUPLICATE ROW {row_ref}: overwriting {prev['mbps']:.2f} Mbps "
                            f"(pg {prev.get('page','?')}: {prev['title'][:50]}) "
                            f"with {max_mbps:.2f} Mbps "
                            f"(pg {page_num}: {g['title'][:50]})"
                        )
                    results[row_ref] = {
                        "mbps": max_mbps,
                        "in_mbps": in_max,
                        "out_mbps": out_max,
                        "title": g["title"],
                        "desc": desc,
                        "page": page_num,
                        "suspect": g.get("suspect", False),
                        "in_suspect": g.get("in_suspect", False),
                        "out_suspect": g.get("out_suspect", False),
                        "extraction_failed": extraction_failed,
                    }
                log.info(f"  MATCH: {desc} -> {row_ref} = {max_mbps:.2f} Mbps")
            else:
                unmatched.append(info)
                log.warning(f"  UNMATCHED: {g['title'][:70]}")

    out = {
        "results": results,
        "unmatched": unmatched,
        "could_not_open": could_not_open,
        "all_graphs": all_graphs,
        "total_pages": total_pages,
    }
    if collect_raw_ocr:
        out["raw_ocr_pages"] = raw_ocr_pages
    return out


# =====================================================================
# SECTION 4 — XLSX GENERATION
# =====================================================================

def _correct_value_pair(in_mbps: float, out_mbps: float, allocated: float,
                        in_suspect: bool, out_suspect: bool) -> tuple:
    """
    Auto-correct OCR decimal-drop errors using allocated bandwidth as sanity ceiling.

    Two failure modes handled:
      HIGH: OCR drops decimal in a G/M value  e.g. "2.93G" read as "293G" = 293,000 Mbps
            Fix: divide by 10/100/1000 until value <= allocated
      LOW:  OCR drops decimal AND unit (unitless bps-rule applied)
            e.g. "37.294M" read as "37294" (no unit) -> /1M bps rule -> 0.037 Mbps
            Fix: reinterpret raw OCR number as kbps -> Mbps
            Only applied to the direction that was actually unitless (per-direction suspect).

    Correction is applied independently to inbound and outbound, so the correct
    value from one direction is not lost when the other direction is wrong.

    Returns: (corrected_in, corrected_out, was_corrected: bool)
    """
    if not allocated or allocated <= 0:
        return in_mbps, out_mbps, False

    ceiling = allocated * 1.5   # allow up to 50% burst above allocated
    floor_ratio = 0.005          # < 0.5% of allocated = suspect (only when bps-converted)

    def fix_high(val):
        # AUTO-CORRECT: values > 10x allocated are clearly impossible OCR errors.
        # Values 1.5x–10x: auto-correct if dividing by 10 or 100 yields a plausible
        # value (within allocated), since OCR commonly drops a decimal point
        # (e.g. "909.11 M" read as "90911 M").
        # Values 1.5x–3x without a plausible correction are only highlighted yellow.
        if val is None:
            return val, False
        if val > allocated * 10:
            for div in (10, 100, 1000, 10000):
                candidate = round(val / div, 2)
                if candidate <= allocated:
                    log.info(f"    [AUTO-CORRECT] Decimal-drop (high ÷{div}): {val:.2f} -> {candidate:.2f} Mbps")
                    return candidate, True
            return val, False
        if val > ceiling:
            # 1.5x–10x range: auto-correct if ÷10 or ÷100 gives a plausible value
            for div in (10, 100):
                candidate = round(val / div, 2)
                if 0 < candidate <= allocated:
                    log.info(
                        f"    [AUTO-CORRECT] Decimal-drop (mid ÷{div}): {val:.2f} -> {candidate:.2f} Mbps "
                        f"({val/allocated:.1f}x allocated)"
                    )
                    return candidate, True
            # No plausible correction found — warn but don't correct
            log.warning(
                f"    [OCR-SUSPECT] {val:.2f} Mbps is {val/allocated:.1f}x allocated — "
                f"verify manually"
            )
        return val, False

    def fix_low(val, is_suspect):
        if val is None or not is_suspect or allocated <= 10:
            return val, False
        if val >= allocated * floor_ratio:
            return val, False
        # Reverse the /1M bps rule to recover the original OCR integer.
        raw = val * 1_000_000
        # Use a tighter ceiling for low-value recovery: the recovered value should
        # be well within allocated bandwidth (≤ allocated), not at burst levels.
        # A near-zero value "recovered" to 144% of allocation is almost certainly
        # a false positive (legitimate low traffic, not an OCR unit-drop).
        low_ceiling = allocated
        # Attempt 1: raw is already Mbps (OCR dropped the 'M' unit but kept correct digits)
        # e.g. "751 M" -> OCR reads "751" -> bps -> raw=751 -> direct Mbps: 751
        if 0 < raw <= low_ceiling:
            log.info(f"    [AUTO-CORRECT] Decimal-drop (low): {val:.6f} -> {raw:.3f} Mbps (direct Mbps)")
            return round(raw, 3), True
        # Attempt 2: raw is in kbps (OCR dropped decimal, e.g. "37.294 M" -> "37294" -> bps -> raw=37294 -> /1000=37.294)
        candidate = round(raw / 1000, 3)
        if 0 < candidate <= low_ceiling:
            log.info(f"    [AUTO-CORRECT] Decimal-drop (low): {val:.6f} -> {candidate:.3f} Mbps (kbps reinterpret)")
            return candidate, True
        return val, False

    new_in,  fixed_in  = fix_high(in_mbps)
    new_out, fixed_out = fix_high(out_mbps)

    # Revert a single-direction fix_high when the corrected value falls below the
    # other direction's raw value — indicates a false positive (legitimate traffic
    # burst rather than an OCR decimal-drop error, e.g. 167.53M on a ~15M allocated
    # link would be reduced to 16.75M, then max(16.75, 69.97)=69.97 selected).
    if fixed_in and not fixed_out and new_in is not None and out_mbps is not None and new_in < out_mbps:
        new_in, fixed_in = in_mbps, False
    elif fixed_out and not fixed_in and new_out is not None and in_mbps is not None and new_out < in_mbps:
        new_out, fixed_out = out_mbps, False

    if not fixed_in:
        new_in,  fixed_in  = fix_low(in_mbps,  in_suspect)
    if not fixed_out:
        new_out, fixed_out = fix_low(out_mbps, out_suspect)

    # Revert fix_low when one direction was "recovered" but the other direction
    # is trustworthy — either it had a valid unit (not suspect) or it has a
    # normal reading above the floor. Indicates genuine low traffic, not OCR error.
    if fixed_in and not fixed_out:
        if not out_suspect or (out_mbps or 0) >= allocated * floor_ratio:
            new_in, fixed_in = in_mbps, False
    elif fixed_out and not fixed_in:
        if not in_suspect or (in_mbps or 0) >= allocated * floor_ratio:
            new_out, fixed_out = out_mbps, False

    # If BOTH directions got fix_low'd, revert both — two simultaneous
    # OCR unit drops is extremely unlikely; this pattern almost always
    # indicates an idle link with legitimately tiny values.
    if fixed_in and fixed_out:
        new_in, fixed_in = in_mbps, False
        new_out, fixed_out = out_mbps, False

    return new_in, new_out, (fixed_in or fixed_out)


# ---------------------------------------------------------------------------
# Fill / font / border constants  (8-char ARGB = fully opaque)
# ---------------------------------------------------------------------------

# Traffic-light fills for E column (utilization-based)
_FILL_GREEN  = PatternFill(start_color="FF92D050", fill_type="solid")  # ≤70%  healthy
_FILL_AMBER  = PatternFill(start_color="FFFFC000", fill_type="solid")  # 71–90% caution
_FILL_ORANGE = PatternFill(start_color="FFFF6600", fill_type="solid")  # 91–100% warning
_FILL_RED    = PatternFill(start_color="FFFF0000", fill_type="solid")  # >100% exceeded
_FILL_BLUE   = PatternFill(start_color="FF9DC3E6", fill_type="solid")  # OCR auto-corrected

# F column — unmatched row (no graph found for this client)
_FILL_YELLOW = PatternFill(start_color="FFFFFF00", fill_type="solid")  # opaque yellow

# Section header rows (rows 3, 30, 53, 57)
_FILL_HEADER = PatternFill(start_color="FF375623", fill_type="solid")  # dark green
_FONT_HEADER = Font(name="Calibri", size=11, bold=True, color="FFFFFFFF")  # white bold

# Cell borders
_THIN_SIDE   = Side(style="thin")
_MEDIUM_SIDE = Side(style="medium")
_CELL_BORDER  = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_THIN_SIDE)
_TITLE_BORDER = Border(left=_THIN_SIDE, right=_THIN_SIDE, top=_THIN_SIDE, bottom=_MEDIUM_SIDE)


def _pick_e_fill(pct: float, corrected: bool):
    """
    Return the traffic-light fill for an E-column cell.

    Args:
        pct:       (mbps / allocated_mbps) * 100  — utilization percentage
        corrected: True if the value was auto-corrected by the OCR decimal-drop fix

    Priority:
        > 100% → always red (over allocation), regardless of correction
        91–100% → orange
        71–90%  → amber
        ≤ 70% + corrected → blue  (within safe range but OCR-touched, flag for awareness)
        ≤ 70%             → green
    """
    if pct > 100:
        return _FILL_RED
    if pct > 90:
        return _FILL_ORANGE
    if pct > 70:
        return _FILL_AMBER
    if corrected:
        return _FILL_BLUE
    return _FILL_GREEN


# Section header rows — these rows get dark-green fill + white bold font
SECTION_HEADER_ROWS = {3, 30, 53, 57}


# All E-column rows expected to be filled (used to detect unfilled rows)
# NOTE: E40 and E51 excluded — no MRTG graph patterns target these rows.
# Re-add them here if mapping patterns are added to GRAPH_TO_ROW_MAP.
EXPECTED_E_ROWS = (
    [f"E{i}" for i in range(4, 29)] +
    [f"E{i}" for i in range(31, 40)] +
    [f"E{i}" for i in range(41, 51)] +
    ["E54", "E55"] +
    [f"E{i}" for i in range(58, 68)]
)


def _fix_apply_fill(output_path: str):
    """
    openpyxl omits applyFill="1" from xf entries even when a non-default fill is
    set.  Excel requires this attribute to display cell fills — without it the
    fillId is silently ignored and cells appear unfilled.

    This post-processes styles.xml inside the saved xlsx zip to add the missing
    attribute to every xf entry whose fillId is non-zero.
    """
    import zipfile as _zf
    tmp = output_path + "._patch.tmp"
    try:
        with _zf.ZipFile(output_path, "r") as zin, \
             _zf.ZipFile(tmp, "w", _zf.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == "xl/styles.xml":
                    xml = data.decode("utf-8")

                    def _patch(m):
                        tag = m.group(0)
                        fm = re.search(r'fillId="(\d+)"', tag)
                        if fm and fm.group(1) != "0" and "applyFill" not in tag:
                            tag = tag.replace("fillId=", 'applyFill="1" fillId=', 1)
                        bm = re.search(r'borderId="(\d+)"', tag)
                        if bm and bm.group(1) != "0" and "applyBorder" not in tag:
                            tag = tag.replace("borderId=", 'applyBorder="1" borderId=', 1)
                        return tag

                    xml = re.sub(r"<xf\b[^>]*(?:/>|>)", _patch, xml)
                    data = xml.encode("utf-8")
                zout.writestr(item, data)
        os.replace(tmp, output_path)
        log.info("  Style patch applied (applyFill='1' and applyBorder='1' added to styles.xml)")
    except Exception as exc:
        log.warning(f"  Style patch failed — fills/borders may not display in Excel: {exc}")
        if os.path.exists(tmp):
            os.remove(tmp)


def generate_report(template_path: str, extraction_data: dict, output_path: str, report_date: str = None):
    """
    Load template xlsx, fill in E-column values, update title date, save output.

    Highlighting rules:
    - E cell -> yellow if extracted value exceeds column D allocated bandwidth
    - E cell -> yellow if value was extracted without a unit (suspect bps conversion)
      and the result is >= 1.0 Mbps (suspicious after bps conversion)
    - E cell -> orange if value was auto-corrected by decimal-drop algorithm (review advised)
    - F cell -> yellow if the row was not filled at all (unmatched/missing graph)
    - E56 -> always set to formula =SUM(E54,E55) (Cache Total)
    """
    log.info(f"Loading template: {template_path}")
    wb = load_workbook(template_path)
    ws = wb.active
    # (fills are module-level constants — no init call needed)

    if report_date:
        ws["A1"] = f"Daily Usage Report ({report_date})"

    filled_rows = set()
    filled = 0

    for row_ref, data in extraction_data.items():
        mbps = data["mbps"]
        suspect = data.get("suspect", False)
        try:
            cell = ws[row_ref]
            row_num = int(re.search(r"\d+", row_ref).group())

            # --- Read allocated bandwidth (column D) — skip formula cells ---
            d_raw = ws[f"D{row_num}"].value
            d_val = d_raw if isinstance(d_raw, (int, float)) else None

            # --- Auto-correct decimal-drop OCR errors using allocated bandwidth ---
            corrected = False
            if d_val and d_val > 0:
                in_mbps  = data.get("in_mbps",  mbps) or 0.0
                out_mbps = data.get("out_mbps", mbps) or 0.0
                new_in, new_out, corrected = _correct_value_pair(
                    in_mbps, out_mbps, d_val,
                    data.get("in_suspect", suspect),
                    data.get("out_suspect", suspect),
                )
                if corrected:
                    mbps = max(
                        (v for v in [new_in, new_out] if v is not None and v >= 0),
                        default=0.0,
                    )

            # --- Write value ---
            if mbps is not None and mbps > 0:
                cell.value = round(mbps, 2) if mbps < 100 else round(mbps)
            elif mbps == 0:
                cell.value = 0

            filled_rows.add(row_ref)
            filled += 1

            # --- Traffic-light fill ---
            if d_val and d_val > 0 and mbps is not None and mbps >= 0:
                pct = (mbps / d_val) * 100
                cell.fill = _pick_e_fill(pct, corrected)
                fill_label = f"{pct:.0f}%"
            else:
                fill_label = "no-alloc"

            # --- Number format: integer for ≥100 Mbps, 2dp for smaller values ---
            if isinstance(cell.value, (int, float)) and cell.value is not None:
                cell.number_format = "#,##0" if cell.value >= 100 else "#,##0.00"

            log.info(f"  {row_ref} = {cell.value}  ({data.get('desc', '')}) [{fill_label}]")
        except Exception as e:
            log.error(f"  Failed to write {row_ref}: {e}")

    # --- Highlight F cell yellow for rows that were NOT filled ---
    for row_ref in EXPECTED_E_ROWS:
        if row_ref not in filled_rows:
            try:
                row_num = int(re.search(r"\d+", row_ref).group())
                f_cell = ws[f"F{row_num}"]
                f_cell.fill = _FILL_YELLOW
                log.info(f"  F{row_num} highlighted yellow (no data for {row_ref})")
            except Exception:
                pass

    # --- Section header rows: dark green fill + white bold text ---
    for row_num in SECTION_HEADER_ROWS:
        for col_letter in ("A", "B", "C", "D", "E", "F"):
            hdr_cell = ws[f"{col_letter}{row_num}"]
            hdr_cell.fill = _FILL_HEADER
            hdr_cell.font = _FONT_HEADER

    # --- Apply borders to all data cells (A:F, rows 1 to last used row) ---
    _DATA_COLS = ("A", "B", "C", "D", "E", "F")
    max_row = max(ws.max_row, 68)
    for row_num in range(1, max_row + 1):
        border = _TITLE_BORDER if row_num == 1 else _CELL_BORDER
        for col_letter in _DATA_COLS:
            ws[f"{col_letter}{row_num}"].border = border

    # --- Fix Cache Total: always write formula (not static value) ---
    ws["E56"] = "=SUM(E54,E55)"
    log.info("  E56 set to =SUM(E54,E55) (Cache Total formula)")

    wb.save(output_path)
    _fix_apply_fill(output_path)
    log.info(f"Report saved: {output_path} ({filled} cells filled)")
    return output_path


# =====================================================================
# SECTION 5 — CLI MODE
# =====================================================================

_MONTH_NAMES = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]


def _extract_date_from_filename(filename: str) -> str:
    """
    Try to extract a human-readable date from a PDF filename.

    Patterns tried (examples):
      26_March_2026.pdf  ->  "26 March 2026"
      26-March-2026.pdf  ->  "26 March 2026"
      2026-03-26.pdf     ->  "26 March 2026"
      20260326.pdf       ->  "26 March 2026"
    Falls back to today's date if no pattern matches.
    """
    base = os.path.splitext(os.path.basename(filename))[0]

    # "26_March_2026" or "26-March-2026" or "26 March 2026"
    m = re.search(
        r"(\d{1,2})[_\-\s](" + "|".join(_MONTH_NAMES) + r")[_\-\s](\d{4})",
        base, re.I,
    )
    if m:
        day, month, year = m.group(1), m.group(2).capitalize(), m.group(3)
        return f"{int(day)} {month} {year}"

    # "2026-03-26" or "26-03-2026"
    m = re.search(r"(\d{4})[_\-](\d{2})[_\-](\d{2})", base)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 1 <= mo <= 12:
            return f"{d} {_MONTH_NAMES[mo - 1]} {y}"

    m = re.search(r"(\d{2})[_\-](\d{2})[_\-](\d{4})", base)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 1 <= mo <= 12:
            return f"{d} {_MONTH_NAMES[mo - 1]} {y}"

    # "20260326"
    m = re.search(r"(\d{4})(\d{2})(\d{2})", base)
    if m:
        y, mo, d = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 1 <= mo <= 12 and 1 <= d <= 31:
            return f"{d} {_MONTH_NAMES[mo - 1]} {y}"

    return datetime.now().strftime("%d %B %Y")


def _print_summary(data: dict):
    print(f"\n{'='*50}")
    print(f"Extraction Summary:")
    print(f"  Pages:           {data['total_pages']}")
    print(f"  Graphs matched:  {len(data['results'])}")
    print(f"  Unmatched:       {len(data['unmatched'])}")
    print(f"  Could not open:  {data['could_not_open']}")

    if data["unmatched"]:
        print(f"\nUnmatched graphs (may need mapping updates):")
        for g in data["unmatched"]:
            print(f"  Page {g['page']}: {g['title'][:70]}  (max={g['max_mbps']:.2f})")

    print(f"\nMatched values:")
    for ref in sorted(data["results"].keys(), key=lambda x: (x[0], int(re.search(r'\d+', x).group()))):
        info = data["results"][ref]
        print(f"  {ref} = {info['mbps']:.2f} Mbps  ({info['desc']})")


def _save_debug_json(data: dict, path: str, full: bool = False):
    debug = {
        "results": {k: {"mbps": v["mbps"], "desc": v["desc"]} for k, v in data["results"].items()},
        "unmatched": [{"page": g["page"], "title": g["title"], "max_mbps": g["max_mbps"]} for g in data["unmatched"]],
        "could_not_open": data["could_not_open"],
    }
    if full and "raw_ocr_pages" in data:
        debug["raw_ocr_pages"] = data["raw_ocr_pages"]
    with open(path, "w", encoding="utf-8") as f:
        json.dump(debug, f, indent=2, ensure_ascii=False)
    print(f"Debug JSON: {path}")


def run_cli():
    parser = argparse.ArgumentParser(description="MRTG Bandwidth Report Generator (CLI)")
    # Single-file mode
    parser.add_argument("--pdf", help="Input PDF with MRTG graphs (single-file mode)")
    parser.add_argument("--output", help="Output xlsx path (auto-generated if omitted)")
    # Batch mode
    parser.add_argument("--batch", metavar="DIR",
                        help="Batch mode: process all PDFs in DIR, auto-name outputs by date")
    parser.add_argument("--output-dir", metavar="DIR",
                        help="Batch mode: directory for output xlsx files (defaults to same dir as each PDF)")
    # Shared
    parser.add_argument("--template", required=True, help="Template xlsx file")
    parser.add_argument("--date", help="Report date, e.g. '26 March 2026' (single-file; auto-detected in batch)")
    parser.add_argument("--dpi", type=int, default=250, help="PDF render DPI (default: 250)")
    parser.add_argument("--warn-duplicates", action="store_true",
                        help="Warn when two graphs map to the same row (shows both values)")
    parser.add_argument("--debug-json", help="Save debug info to JSON file (single-file mode)")
    parser.add_argument("--debug-full", action="store_true",
                        help="Include raw per-page OCR text in debug JSON (requires --debug-json)")
    args = parser.parse_args()

    if not os.path.isfile(args.template):
        print(f"ERROR: Template not found: {args.template}"); sys.exit(1)

    # ── Batch mode ──────────────────────────────────────────────────
    if args.batch:
        if not os.path.isdir(args.batch):
            print(f"ERROR: Batch directory not found: {args.batch}"); sys.exit(1)
        pdfs = sorted(
            p for p in (os.path.join(args.batch, f) for f in os.listdir(args.batch))
            if p.lower().endswith(".pdf") and os.path.isfile(p)
        )
        if not pdfs:
            print(f"ERROR: No PDF files found in {args.batch}"); sys.exit(1)

        out_dir = args.output_dir or None
        if out_dir and not os.path.isdir(out_dir):
            os.makedirs(out_dir, exist_ok=True)

        print(f"Batch mode: {len(pdfs)} PDF(s) found in {args.batch}")
        for pdf_path in pdfs:
            date_str = _extract_date_from_filename(pdf_path)
            dest_dir = out_dir or os.path.dirname(pdf_path)
            out_path = os.path.join(dest_dir, f"Bandwidth Report (MAX) For {date_str}.xlsx")
            print(f"\n[BATCH] {os.path.basename(pdf_path)}  ->  date={date_str}")
            try:
                data = extract_all_graphs(
                    pdf_path, dpi=args.dpi,
                    warn_duplicates=args.warn_duplicates,
                    collect_raw_ocr=args.debug_full,
                )
                _print_summary(data)
                generate_report(args.template, data["results"], out_path, date_str)
                print(f"  Output: {out_path}")
            except Exception as exc:
                print(f"  ERROR processing {pdf_path}: {exc}")
        return

    # ── Single-file mode ─────────────────────────────────────────────
    if not args.pdf:
        print("ERROR: --pdf is required in single-file mode (or use --batch DIR)"); sys.exit(1)
    if not os.path.isfile(args.pdf):
        print(f"ERROR: PDF not found: {args.pdf}"); sys.exit(1)

    if not args.output:
        date_str = args.date or datetime.now().strftime("%d %B %Y")
        args.output = f"Bandwidth Report (MAX) For {date_str}.xlsx"

    print(f"Extracting from: {args.pdf}")
    data = extract_all_graphs(
        args.pdf, dpi=args.dpi,
        warn_duplicates=args.warn_duplicates,
        collect_raw_ocr=args.debug_full,
    )
    _print_summary(data)
    generate_report(args.template, data["results"], args.output, args.date)
    print(f"\nOutput: {args.output}")

    if args.debug_json:
        _save_debug_json(data, args.debug_json, full=args.debug_full)


# =====================================================================
# SECTION 6 — GUI (tkinter)
# =====================================================================

_CONFIG_PATH = os.path.join(os.path.expanduser("~"), ".mrtg_report_config.json")


def _load_gui_config() -> dict:
    """Load persisted GUI settings from ~/.mrtg_report_config.json."""
    try:
        with open(_CONFIG_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}


def _save_gui_config(cfg: dict):
    """Persist GUI settings to ~/.mrtg_report_config.json."""
    try:
        with open(_CONFIG_PATH, "w", encoding="utf-8") as f:
            json.dump(cfg, f, indent=2)
    except Exception as exc:
        log.warning(f"Could not save GUI config: {exc}")


def run_gui():
    import tkinter as tk
    from tkinter import ttk, filedialog, messagebox, scrolledtext
    import threading

    class MRTGApp:
        def __init__(self, root):
            self.root = root
            self.root.title("MRTG Bandwidth Report Generator")
            self.root.geometry("880x760")
            self.root.minsize(720, 580)

            # Shared state
            cfg = _load_gui_config()
            self.template_path  = tk.StringVar(value=cfg.get("template_path", ""))
            self.dpi            = tk.IntVar(value=cfg.get("dpi", 250))
            self.warn_dups      = tk.BooleanVar(value=cfg.get("warn_duplicates", False))
            self.running        = False

            # Single-file state
            self.pdf_path       = tk.StringVar(value=cfg.get("pdf_path", ""))
            self.output_path    = tk.StringVar()
            self.report_date    = tk.StringVar(value=datetime.now().strftime("%d %B %Y"))
            self._output_manually_set = False
            self.report_date.trace_add("write", self._on_date_changed)
            self._last_unfilled = []

            # Batch state
            self.batch_dir      = tk.StringVar(value=cfg.get("batch_dir", ""))
            self.batch_out_dir  = tk.StringVar(value=cfg.get("batch_out_dir", ""))

            self._build_ui()
            # Restore auto output path after UI is built
            if self.pdf_path.get():
                self.output_path.set(self._auto_output_path())

        # ── Config persistence ─────────────────────────────────────────

        def _persist_config(self):
            _save_gui_config({
                "pdf_path":       self.pdf_path.get(),
                "template_path":  self.template_path.get(),
                "dpi":            self.dpi.get(),
                "warn_duplicates": self.warn_dups.get(),
                "batch_dir":      self.batch_dir.get(),
                "batch_out_dir":  self.batch_out_dir.get(),
            })

        # ── UI construction ────────────────────────────────────────────

        def _build_ui(self):
            outer = ttk.Frame(self.root, padding=10)
            outer.pack(fill=tk.BOTH, expand=True)

            ttk.Label(outer, text="MRTG Bandwidth Report Generator",
                      font=("Helvetica", 15, "bold")).pack(pady=(0, 8))

            # Shared: template + DPI + options
            shared = ttk.LabelFrame(outer, text="Shared Settings", padding=8)
            shared.pack(fill=tk.X, pady=4)

            rs1 = ttk.Frame(shared); rs1.pack(fill=tk.X, pady=2)
            ttk.Label(rs1, text="Template XLSX:", width=16, anchor="e").pack(side=tk.LEFT)
            ttk.Entry(rs1, textvariable=self.template_path, width=52).pack(side=tk.LEFT, padx=4, fill=tk.X, expand=True)
            ttk.Button(rs1, text="Browse...", command=self._browse_template).pack(side=tk.LEFT)

            rs2 = ttk.Frame(shared); rs2.pack(fill=tk.X, pady=2)
            ttk.Label(rs2, text="DPI:", width=16, anchor="e").pack(side=tk.LEFT)
            ttk.Spinbox(rs2, from_=100, to=400, textvariable=self.dpi, width=8).pack(side=tk.LEFT, padx=4)
            ttk.Label(rs2, text="(higher = better OCR, slower)").pack(side=tk.LEFT, padx=6)
            ttk.Checkbutton(rs2, text="Warn on duplicate rows",
                            variable=self.warn_dups,
                            command=self._persist_config).pack(side=tk.LEFT, padx=16)

            # Tab notebook: Single File / Batch
            nb = ttk.Notebook(outer)
            nb.pack(fill=tk.BOTH, expand=True, pady=4)

            self._build_single_tab(nb)
            self._build_batch_tab(nb)

            # Progress (shared)
            pf = ttk.Frame(outer); pf.pack(fill=tk.X, pady=4)
            self.progress_var = tk.DoubleVar()
            ttk.Progressbar(pf, variable=self.progress_var, maximum=100).pack(
                fill=tk.X, side=tk.LEFT, expand=True, padx=(0, 8))
            self.progress_label = ttk.Label(pf, text="Ready", width=32)
            self.progress_label.pack(side=tk.LEFT)

            # Log (shared)
            lf = ttk.LabelFrame(outer, text="Processing Log", padding=4)
            lf.pack(fill=tk.BOTH, expand=True, pady=4)
            self.log_text = scrolledtext.ScrolledText(lf, height=12, font=("Courier", 9))
            self.log_text.pack(fill=tk.BOTH, expand=True)

        def _build_single_tab(self, nb):
            tab = ttk.Frame(nb, padding=8)
            nb.add(tab, text="Single File")

            r1 = ttk.Frame(tab); r1.pack(fill=tk.X, pady=3)
            ttk.Label(r1, text="Input PDF:", width=14, anchor="e").pack(side=tk.LEFT)
            ttk.Entry(r1, textvariable=self.pdf_path, width=52).pack(side=tk.LEFT, padx=4, fill=tk.X, expand=True)
            ttk.Button(r1, text="Browse...", command=self._browse_pdf).pack(side=tk.LEFT)

            r2 = ttk.Frame(tab); r2.pack(fill=tk.X, pady=3)
            ttk.Label(r2, text="Report Date:", width=14, anchor="e").pack(side=tk.LEFT)
            ttk.Entry(r2, textvariable=self.report_date, width=30).pack(side=tk.LEFT, padx=4)

            r3 = ttk.Frame(tab); r3.pack(fill=tk.X, pady=3)
            ttk.Label(r3, text="Output File:", width=14, anchor="e").pack(side=tk.LEFT)
            ttk.Entry(r3, textvariable=self.output_path, width=52).pack(side=tk.LEFT, padx=4, fill=tk.X, expand=True)
            ttk.Button(r3, text="Browse...", command=self._browse_output).pack(side=tk.LEFT)

            bf = ttk.Frame(tab); bf.pack(pady=6)
            self.run_btn = ttk.Button(bf, text="Generate Report", command=self._run_single)
            self.run_btn.pack(side=tk.LEFT, padx=4)
            ttk.Button(bf, text="View Mapping", command=self._show_mapping).pack(side=tk.LEFT, padx=4)
            self.copy_unmatched_btn = ttk.Button(
                bf, text="Copy Unmatched Entries",
                command=self._copy_unmatched, state="disabled")
            self.copy_unmatched_btn.pack(side=tk.LEFT, padx=4)
            ttk.Button(bf, text="Help", command=self._show_help).pack(side=tk.LEFT, padx=4)

        def _build_batch_tab(self, nb):
            tab = ttk.Frame(nb, padding=8)
            nb.add(tab, text="Batch Mode")

            r1 = ttk.Frame(tab); r1.pack(fill=tk.X, pady=3)
            ttk.Label(r1, text="PDFs Directory:", width=16, anchor="e").pack(side=tk.LEFT)
            ttk.Entry(r1, textvariable=self.batch_dir, width=52).pack(side=tk.LEFT, padx=4, fill=tk.X, expand=True)
            ttk.Button(r1, text="Browse...", command=self._browse_batch_dir).pack(side=tk.LEFT)

            r2 = ttk.Frame(tab); r2.pack(fill=tk.X, pady=3)
            ttk.Label(r2, text="Output Dir:", width=16, anchor="e").pack(side=tk.LEFT)
            ttk.Entry(r2, textvariable=self.batch_out_dir, width=52).pack(side=tk.LEFT, padx=4, fill=tk.X, expand=True)
            ttk.Button(r2, text="Browse...", command=self._browse_batch_out_dir).pack(side=tk.LEFT)
            ttk.Label(r2, text="(blank = same dir as each PDF)", foreground="gray").pack(side=tk.LEFT, padx=6)

            ttk.Label(tab,
                text="Dates are auto-detected from each PDF's filename.\n"
                     "Output files are named: Bandwidth Report (MAX) For {date}.xlsx",
                foreground="gray").pack(pady=(4, 0), anchor="w")

            bf = ttk.Frame(tab); bf.pack(pady=8)
            self.batch_run_btn = ttk.Button(bf, text="Run Batch", command=self._run_batch)
            self.batch_run_btn.pack(side=tk.LEFT, padx=4)

        # ── Browse helpers ─────────────────────────────────────────────

        def _auto_output_path(self):
            pdf = self.pdf_path.get()
            if not pdf:
                return ""
            d = self.report_date.get().strip()
            return os.path.join(os.path.dirname(pdf), f"Bandwidth Report (MAX) For {d}.xlsx")

        def _on_date_changed(self, *_):
            if not self._output_manually_set:
                self.output_path.set(self._auto_output_path())

        def _browse_pdf(self):
            p = filedialog.askopenfilename(title="Select Input PDF",
                                           filetypes=[("PDF", "*.pdf"), ("All", "*.*")])
            if p:
                self.pdf_path.set(p)
                self._output_manually_set = False
                self.output_path.set(self._auto_output_path())
                self._persist_config()

        def _browse_template(self):
            p = filedialog.askopenfilename(title="Select Template XLSX",
                                           filetypes=[("Excel", "*.xlsx"), ("All", "*.*")])
            if p:
                self.template_path.set(p)
                self._persist_config()

        def _browse_output(self):
            p = filedialog.asksaveasfilename(title="Save Output As", defaultextension=".xlsx",
                                             filetypes=[("Excel", "*.xlsx"), ("All", "*.*")])
            if p:
                self._output_manually_set = True
                self.output_path.set(p)

        def _browse_batch_dir(self):
            p = filedialog.askdirectory(title="Select directory containing PDFs")
            if p:
                self.batch_dir.set(p)
                self._persist_config()

        def _browse_batch_out_dir(self):
            p = filedialog.askdirectory(title="Select output directory for xlsx files")
            if p:
                self.batch_out_dir.set(p)
                self._persist_config()

        # ── Logging / progress ─────────────────────────────────────────

        def _log(self, msg):
            self.log_text.insert(tk.END, msg + "\n")
            self.log_text.see(tk.END)

        def _progress(self, cur, total, msg):
            self.progress_var.set((cur / total) * 100 if total else 0)
            self.progress_label.config(text=msg)
            self.root.update_idletasks()

        # ── Single-file run ────────────────────────────────────────────

        def _run_single(self):
            if self.running:
                return
            pdf = self.pdf_path.get()
            tpl = self.template_path.get()
            out = self.output_path.get()
            date = self.report_date.get()
            dpi = self.dpi.get()
            warn_dups = self.warn_dups.get()

            if not pdf or not os.path.isfile(pdf):
                messagebox.showerror("Error", "Please select a valid input PDF."); return
            if not tpl or not os.path.isfile(tpl):
                messagebox.showerror("Error", "Please select a valid template XLSX."); return
            if not out:
                messagebox.showerror("Error", "Please set an output file path."); return

            self.running = True
            self.run_btn.config(state="disabled")
            self.log_text.delete("1.0", tk.END)
            self._log(f"PDF: {pdf}")
            self._log(f"Template: {tpl}")
            self._log(f"DPI: {dpi}  |  Warn duplicates: {warn_dups}\n")

            def worker():
                try:
                    data = extract_all_graphs(
                        pdf, dpi=dpi,
                        warn_duplicates=warn_dups,
                        progress_cb=lambda c, t, m: self.root.after(0, self._progress, c, t, m),
                    )
                    self.root.after(0, self._log, "\n--- Extraction Complete ---")
                    self.root.after(0, self._log, f"Matched:  {len(data['results'])}")
                    self.root.after(0, self._log, f"Unmatched: {len(data['unmatched'])}")
                    self.root.after(0, self._log, f"Could not open: {data['could_not_open']}")

                    if data["unmatched"]:
                        self.root.after(0, self._log, "\nUnmatched (may need mapping):")
                        for g in data["unmatched"]:
                            self.root.after(0, self._log,
                                f"  Pg{g['page']}: {g['title'][:65]} (max={g['max_mbps']:.2f})")

                    self.root.after(0, self._log, "\nMatched values:")
                    for ref in sorted(data["results"]):
                        info = data["results"][ref]
                        self.root.after(0, self._log,
                            f"  {ref} = {info['mbps']:.2f} Mbps  ({info['desc']})")

                    self.root.after(0, self._log, "\nGenerating report ...")
                    generate_report(tpl, data["results"], out, date)
                    self.root.after(0, self._log, f"\nSaved: {out}")
                    self.root.after(0, self._progress, 100, 100, "Done!")

                    _row_desc = {row: desc for _, row, desc in GRAPH_TO_ROW_MAP}
                    unfilled = [
                        (ref, _row_desc.get(ref, "—"))
                        for ref in EXPECTED_E_ROWS
                        if ref not in data["results"]
                    ]
                    self._last_unfilled = unfilled
                    self.root.after(0, lambda: self.copy_unmatched_btn.config(
                        state="normal" if unfilled else "disabled"))

                    self.root.after(0, lambda: messagebox.showinfo("Success",
                        f"Report generated!\n\nMatched: {len(data['results'])}\n"
                        f"Unmatched: {len(data['unmatched'])}\nSaved: {out}"))
                except Exception as e:
                    self.root.after(0, self._log, f"\nERROR: {e}")
                    self.root.after(0, lambda: messagebox.showerror("Error", str(e)))
                finally:
                    self.running = False
                    self.root.after(0, lambda: self.run_btn.config(state="normal"))

            threading.Thread(target=worker, daemon=True).start()

        # ── Batch run ──────────────────────────────────────────────────

        def _run_batch(self):
            if self.running:
                return
            batch_dir = self.batch_dir.get()
            tpl = self.template_path.get()
            out_dir = self.batch_out_dir.get() or None
            dpi = self.dpi.get()
            warn_dups = self.warn_dups.get()

            if not batch_dir or not os.path.isdir(batch_dir):
                messagebox.showerror("Error", "Please select a valid PDFs directory."); return
            if not tpl or not os.path.isfile(tpl):
                messagebox.showerror("Error", "Please select a valid template XLSX."); return

            pdfs = sorted(
                p for p in (os.path.join(batch_dir, f) for f in os.listdir(batch_dir))
                if p.lower().endswith(".pdf") and os.path.isfile(p)
            )
            if not pdfs:
                messagebox.showerror("Error", f"No PDF files found in:\n{batch_dir}"); return

            if out_dir and not os.path.isdir(out_dir):
                os.makedirs(out_dir, exist_ok=True)

            self.running = True
            self.batch_run_btn.config(state="disabled")
            self.log_text.delete("1.0", tk.END)
            self._log(f"Batch dir: {batch_dir}")
            self._log(f"Template:  {tpl}")
            self._log(f"Output to: {out_dir or '(same as each PDF)'}")
            self._log(f"DPI: {dpi}  |  PDFs found: {len(pdfs)}\n")

            def batch_worker():
                ok, failed = 0, 0
                for i, pdf_path in enumerate(pdfs):
                    date_str = _extract_date_from_filename(pdf_path)
                    dest_dir = out_dir or os.path.dirname(pdf_path)
                    out_path = os.path.join(dest_dir,
                        f"Bandwidth Report (MAX) For {date_str}.xlsx")
                    self.root.after(0, self._progress, i + 1, len(pdfs),
                                    f"{i+1}/{len(pdfs)}: {os.path.basename(pdf_path)}")
                    self.root.after(0, self._log,
                        f"\n[{i+1}/{len(pdfs)}] {os.path.basename(pdf_path)}  (date: {date_str})")
                    try:
                        data = extract_all_graphs(
                            pdf_path, dpi=dpi, warn_duplicates=warn_dups,
                        )
                        self.root.after(0, self._log,
                            f"  Matched: {len(data['results'])}  "
                            f"Unmatched: {len(data['unmatched'])}  "
                            f"CantOpen: {data['could_not_open']}")
                        generate_report(tpl, data["results"], out_path, date_str)
                        self.root.after(0, self._log, f"  -> {out_path}")
                        ok += 1
                    except Exception as exc:
                        self.root.after(0, self._log, f"  ERROR: {exc}")
                        failed += 1

                self.root.after(0, self._progress, 100, 100, "Batch done!")
                self.root.after(0, self._log,
                    f"\nBatch complete: {ok} succeeded, {failed} failed.")
                self.root.after(0, lambda: messagebox.showinfo(
                    "Batch Complete",
                    f"{ok} report(s) generated.\n{failed} failed.\nCheck the log for details."))
                self.running = False
                self.root.after(0, lambda: self.batch_run_btn.config(state="normal"))

            threading.Thread(target=batch_worker, daemon=True).start()

        # ── Helpers ────────────────────────────────────────────────────

        def _copy_unmatched(self):
            if not self._last_unfilled:
                messagebox.showinfo("Copy Unmatched", "No unmatched entries to copy.")
                return
            lines = ["Row\tClient\tReason"]
            for ref, desc in self._last_unfilled:
                lines.append(f"{ref}\t{desc}\tNo graph in this PDF")
            text = "\n".join(lines)
            self.root.clipboard_clear()
            self.root.clipboard_append(text)
            messagebox.showinfo(
                "Copied",
                f"{len(self._last_unfilled)} unmatched entries copied to clipboard.\n\n"
                "Format: Row | Client | Reason\n"
                "(Tab-separated, paste directly into Excel or Notepad)")

        def _show_mapping(self):
            win = tk.Toplevel(self.root)
            win.title("Graph -> Row Mapping")
            win.geometry("750x500")
            txt = scrolledtext.ScrolledText(win, font=("Courier", 9))
            txt.pack(fill=tk.BOTH, expand=True, padx=8, pady=8)
            txt.insert(tk.END, "# Edit GRAPH_TO_ROW_MAP in the script to customize.\n\n")
            txt.insert(tk.END, f"{'PATTERN':<55} {'CELL':<8} DESCRIPTION\n")
            txt.insert(tk.END, "-" * 100 + "\n")
            for pat, row, desc in GRAPH_TO_ROW_MAP:
                txt.insert(tk.END, f"{pat:<55} {row:<8} {desc}\n")
            txt.insert(tk.END, "\n--- FALLBACK ---\n")
            for pat, row, desc in FALLBACK_MAP:
                txt.insert(tk.END, f"{pat:<55} {row:<8} {desc}\n")

        def _show_help(self):
            messagebox.showinfo("Help",
                "MRTG Bandwidth Report Generator\n\n"
                "Single File tab:\n"
                "  1. Select the input PDF containing MRTG/Cacti graphs\n"
                "  2. Select the template XLSX (previous day's report)\n"
                "  3. Set the report date\n"
                "  4. Click 'Generate Report'\n\n"
                "Batch Mode tab:\n"
                "  1. Select a directory containing multiple PDFs\n"
                "  2. Select the template XLSX\n"
                "  3. Optionally set an output directory\n"
                "  4. Click 'Run Batch' — dates are auto-detected from filenames\n\n"
                "Shared settings (template, DPI, warn-duplicates) apply to both modes.\n"
                "Settings are remembered between sessions.\n\n"
                "CLI mode: python mrtg_bandwidth_report.py --cli --help")

    root = tk.Tk()
    MRTGApp(root)
    root.mainloop()


# =====================================================================
# SECTION 7 — DEPENDENCY CHECK + ENTRY POINT
# =====================================================================

def check_dependencies():
    errors = []
    if MISSING_DEPS:
        errors.append(f"Missing Python packages: {', '.join(MISSING_DEPS)}")
        errors.append("Install with: pip install openpyxl pdf2image pytesseract Pillow")

    if "pytesseract" not in " ".join(MISSING_DEPS):
        try:
            pytesseract.get_tesseract_version()
        except Exception:
            errors.append("Tesseract OCR not found.")
            errors.append("  macOS:   brew install tesseract")
            errors.append("  Ubuntu:  sudo apt install tesseract-ocr")
            errors.append("  Windows: choco install tesseract")

    if "pdf2image" not in " ".join(MISSING_DEPS):
        if not shutil.which("pdftoppm"):
            errors.append("Poppler not found (needed by pdf2image).")
            errors.append("  macOS:   brew install poppler")
            errors.append("  Ubuntu:  sudo apt install poppler-utils")
            errors.append("  Windows: choco install poppler")

    return errors


def main():
    if "--cli" in sys.argv:
        errs = check_dependencies()
        if errs:
            print("DEPENDENCY ERRORS:\n  " + "\n  ".join(errs)); sys.exit(1)
        sys.argv.remove("--cli")
        run_cli()
    else:
        errs = check_dependencies()
        if errs:
            try:
                import tkinter as tk
                from tkinter import messagebox
                r = tk.Tk(); r.withdraw()
                messagebox.showerror("Missing Dependencies", "\n".join(errs))
                r.destroy()
            except Exception:
                print("DEPENDENCY ERRORS:\n  " + "\n  ".join(errs))
            sys.exit(1)
        run_gui()


if __name__ == "__main__":
    main()
